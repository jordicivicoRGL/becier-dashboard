# -*- coding: utf-8 -*-
"""Genera la plantilla de biblioteca de eventos de tracking (Excel).

Pestaña "Plantilla": base reutilizable en blanco para cualquier cliente.
Pestaña "DCORE": inventario real (2026-08-24) de GA4 + GTM + Google Ads.
Pestaña "Diagnóstico DCORE": duplicidades y huecos detectados.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

OUT_PATH = r"C:\Users\PC\Desktop\Curso Claude Code VSCode\claudito\clients\_referencias\plantilla_eventos_tracking.xlsx"

HEADER_FILL = PatternFill(start_color="1C4587", end_color="1C4587", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=11)
WARN_FILL = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
OK_FILL = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ["A", "B", "C", "D", "E", "F", "G"]
WIDTHS = {"A": 24, "B": 9, "C": 9, "D": 9, "E": 48, "F": 16, "G": 42}


def write_header(ws: Worksheet, row):
    headers = ["Evento", "Meta", "Google", "TikTok", "Medición", "Fuente", "Notas"]
    for col, text in zip(COLS, headers):
        cell = ws[f"{col}{row}"]
        cell.value = text
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_event_row(ws: Worksheet, row, name, meta, google, tiktok, medicion, fuente, notas="", fill=None):
    ws[f"A{row}"] = name
    ws[f"B{row}"] = "X" if meta else ""
    ws[f"C{row}"] = "X" if google else ""
    ws[f"D{row}"] = "X" if tiktok else ""
    ws[f"E{row}"] = medicion
    ws[f"F{row}"] = fuente
    ws[f"G{row}"] = notas
    for col in COLS:
        cell = ws[f"{col}{row}"]
        cell.border = BORDER
        cell.alignment = CENTER if col in ("B", "C", "D", "F") else LEFT
        if fill:
            cell.fill = fill


def write_section_title(ws: Worksheet, row, title):
    ws[f"A{row}"] = title
    ws[f"A{row}"].font = SECTION_FONT


def apply_widths(ws: Worksheet):
    for col, width in WIDTHS.items():
        ws.column_dimensions[col].width = width


# ---------------------------------------------------------------- Plantilla
def build_template_sheet(ws: Worksheet):
    write_section_title(ws, 2, "Eventos comunes (Meta / Google / TikTok)")
    write_header(ws, 3)
    write_event_row(ws, 4, "Page View", True, True, True, "Mide cuando el usuario llega a una página", "GA4")
    write_event_row(ws, 5, "Lead", True, True, True, "Mide cuando el usuario rellena un formulario", "GTM")
    write_event_row(ws, 6, "", False, False, False, "", "")
    write_event_row(ws, 7, "", False, False, False, "", "")

    write_section_title(ws, 9, "Analytics (GA4 nativo)")
    write_header(ws, 10)
    write_event_row(ws, 11, "form_start", True, False, False,
                     "Evento automático de GA4 (medición mejorada). Salta al interactuar con un formulario. Una vez por sesión.", "GA4")
    write_event_row(ws, 12, "form_submit", True, False, False,
                     "Evento automático de GA4 (medición mejorada). Se dispara al enviarse un formulario.", "GA4")
    write_event_row(ws, 13, "", False, False, False, "", "")

    write_section_title(ws, 15, "Eventos custom (GTM)")
    write_header(ws, 16)
    write_event_row(ws, 17, "", False, False, False, "", "")
    write_event_row(ws, 18, "", False, False, False, "", "")
    write_event_row(ws, 19, "", False, False, False, "", "")

    write_section_title(ws, 21, "Conversiones en Google Ads")
    write_header(ws, 22)
    write_event_row(ws, 23, "", False, False, False, "", "")
    write_event_row(ws, 24, "", False, False, False, "", "")

    write_section_title(ws, 26, "Píxel / eventos Meta")
    write_header(ws, 27)
    write_event_row(ws, 28, "", False, False, False, "", "")
    write_event_row(ws, 29, "", False, False, False, "", "")

    apply_widths(ws)
    ws.freeze_panes = "A4"


# -------------------------------------------------------------------- DCORE
def build_dcore_sheet(ws: Worksheet):
    write_section_title(ws, 2, "GA4 — eventos automáticos (medición mejorada), propiedad dcore.es (G-LKG6KEHBKP)")
    write_header(ws, 3)
    write_event_row(ws, 4, "page_view", False, True, False,
                     "Carga de página. Automático.", "GA4 nativo", "7.671 eventos (27 jul-23 ago)")
    write_event_row(ws, 5, "session_start", False, True, False,
                     "Inicio de sesión. Automático.", "GA4 nativo", "2.597 eventos")
    write_event_row(ws, 6, "first_visit", False, True, False,
                     "Primera visita del usuario. Automático.", "GA4 nativo", "1.936 eventos")
    write_event_row(ws, 7, "user_engagement", False, True, False,
                     "Tiempo de interacción con la página. Automático.", "GA4 nativo", "2.275 eventos")
    write_event_row(ws, 8, "scroll", False, True, False,
                     "Scroll >90% de la página. Automático.", "GA4 nativo", "501 eventos")
    write_event_row(ws, 9, "click", False, True, False,
                     "Clic en enlace saliente. Automático.", "GA4 nativo", "49 eventos")
    write_event_row(ws, 10, "form_start", False, True, False,
                     "Interacción con un formulario. Una vez por sesión. EVENTO CLAVE.", "GA4 nativo",
                     "499 eventos / 388 usuarios. Marcado como evento clave en GA4.", fill=OK_FILL)
    write_event_row(ws, 11, "form_submit", False, True, False,
                     "Envío de un formulario que pasa la validación básica de GA4. EVENTO CLAVE.", "GA4 nativo",
                     "330 eventos / 120 usuarios. Marcado como evento clave en GA4. Importado a Ads pero OCULTO "
                     "('dcore.es (web) form_submit', HIDDEN, primary=False) → no se usa en optimización.", fill=WARN_FILL)
    write_event_row(ws, 12, "purchase", False, True, False,
                     "Evento estándar de ecommerce. Definido pero sin datos (DCORE no vende online).", "GA4 nativo",
                     "0 eventos. No es evento clave. Importado a Ads pero oculto y sin uso.")

    write_section_title(ws, 14, "GTM (GTM-TG55B388) — eventos custom hacia GA4 / plataformas")
    write_header(ws, 15)
    write_event_row(ws, 16, "Form Send", False, True, False,
                     "Tag 'Form Send 2 GA' (GA4 evento). Trigger 'Form Send' = Envío de formulario, SIN filtro "
                     "(cualquier formulario, sin esperar validación). Por eso cuenta más que form_submit.",
                     "GTM → GA4", "442 eventos / 203 usuarios. NO es evento clave en GA4 ahora mismo, pero SÍ está "
                     "activa como conversión Principal en Google Ads (163,50 conv en PMAX). Tag modificada hace 4h "
                     "(revisar qué cambió).", fill=WARN_FILL)
    write_event_row(ws, 17, "GA PV (page_view)", False, True, False,
                     "Tag 'GA PV', evento GA4 manual disparado por trigger 'PageView' propio de GTM.",
                     "GTM → GA4", "Duplica funcionalmente el page_view automático nativo de GA4 (fila 4). Revisar si "
                     "sigue siendo necesario.", fill=WARN_FILL)
    write_event_row(ws, 18, "Conversión Google Ads (gtag)", False, True, False,
                     "Tag 'Form conversion G.ads 2'. Conversion ID 10814118154 / label sUeRCJOtnb4ZEIqyyaQo. "
                     "Trigger 'Thank You Page' = vista de página con URL que CONTIENE 'gracias'.",
                     "GTM → Ads (directo)", "Coincide con /gracias/ Y /lp15/gracias a la vez (lp15/gracias redirige "
                     "a gracias/) → riesgo de disparo doble en la misma sesión.", fill=WARN_FILL)
    write_event_row(ws, 19, "Linker G.Ads", False, True, False,
                     "Tag de vinculación de conversiones (gclid). Mismo trigger 'Thank You Page'.",
                     "GTM", "No es un evento de conversión en sí, solo enlaza clics con conversiones.")
    write_event_row(ws, 20, "FB Pixel / Facebook Conversion Linker", True, False, False,
                     "Tags de Meta Pixel. Triggers 'All Pages' y 'Form Send FB'.",
                     "GTM → Meta", "Aparecen PAUSADAS en GTM (icono de pausa). Coherente con que Meta Ads no tiene "
                     "tráfico activo a la web ahora mismo (todo Lead Ads nativo).")

    write_section_title(ws, 22, "Google Ads — acciones de conversión (cuenta 182-915-0362)")
    write_header(ws, 23)
    write_event_row(ws, 24, "Form send", False, True, False,
                     "Importada desde GA4 (evento custom 'Form Send'). ENABLED, Primary.",
                     "GA4→Ads (import)", "163,50 conv en PMAX (1-23 ago). Basada en un evento que ya NO es clave en "
                     "GA4 → conversión huérfana, revisar si sigue recibiendo datos nuevos.", fill=WARN_FILL)
    write_event_row(ws, 25, "Envío de formulario... (lp15/gracias)", False, True, False,
                     "Auto-detectada por Google Ads (webpage_codeless) a partir del Google Tag base. ENABLED, Primary.",
                     "Ads (auto, page load)", "10,00 conv. Landing antigua en desuso que redirige a /gracias/. "
                     "Candidata a eliminar.", fill=WARN_FILL)
    write_event_row(ws, 26, "Registro (gracias/)", False, True, False,
                     "Auto-detectada por Google Ads (webpage_codeless). ENABLED, Primary.",
                     "Ads (auto, page load)", "3,00 conv. Página de gracias actual. Posible doble conteo con la "
                     "anterior si ambas cargan en la misma sesión.", fill=WARN_FILL)
    write_event_row(ws, 27, "Lead (Thank you page)", False, True, False,
                     "Auto-detectada por Google Ads (webpage_codeless), categoría SIGNUP. ENABLED, Primary.",
                     "Ads (auto, page load)", "Otra conversión activa de tipo lead a nivel de cuenta, no vista aún "
                     "a nivel de campaña PMAX. Revisar solapamiento con las anteriores.", fill=WARN_FILL)
    write_event_row(ws, 28, "Formulario de contacto - Enviar", False, True, False,
                     "Lead Ads nativo de Google (formulario dentro del propio anuncio, sin salir a la web). "
                     "ENABLED, Primary.", "Google Hosted", "No relacionado con la web ni con GA4/GTM.")
    write_event_row(ws, 29, "Calls from ads", False, True, False,
                     "Llamadas iniciadas desde el anuncio.", "Ads nativo", "ENABLED, Primary.")
    write_event_row(ws, 30, "dcore.es (web) form_submit [OCULTA]", False, True, False,
                     "Importación automática desde GA4 del evento clave form_submit.",
                     "GA4→Ads (import)", "HIDDEN, primary=False. Es la más fiable de todas pero no se usa para nada "
                     "en la cuenta ahora mismo.", fill=WARN_FILL)
    write_event_row(ws, 31, "6 acciones eliminadas (histórico)", False, True, False,
                     "Formulario, Lead, Lead (1), Lead (2), Registro Gracias, Envío de formulario (gracias, "
                     "versión antigua).", "—", "REMOVED. No afectan a las métricas actuales pero ensucian la "
                     "vista de la cuenta.")

    apply_widths(ws)
    ws.freeze_panes = "A4"


# ------------------------------------------------------- Diagnóstico DCORE
def build_diagnostico_sheet(ws: Worksheet):
    ws["A1"] = "Diagnóstico de tracking DCORE — 2026-08-24"
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "1. Triple medición del mismo \"envío de formulario\", sin que ninguna sea la fuente única de verdad."
    ws["A3"].font = SECTION_FONT
    ws["A4"] = ("form_start (499) mide interacción, no envío. form_submit (330, evento clave de GA4, pero su "
                "importación a Ads está OCULTA y sin usar) vs Form Send (442, custom vía GTM, SÍ es la conversión "
                "Principal activa en Ads) no coinciden porque Form Send no exige que el formulario pase "
                "validación y form_submit sí.")
    ws["A4"].alignment = LEFT
    ws.row_dimensions[4].height = 45

    ws["A6"] = "2. El trigger de conversión de Google Ads en GTM puede disparar dos veces por sesión."
    ws["A6"].font = SECTION_FONT
    ws["A7"] = ("El trigger 'Thank You Page' (usado por 'Form conversion G.ads 2' y 'Linker G.Ads') se activa con "
                "cualquier URL que contenga 'gracias'. Como /lp15/gracias redirige a /gracias/, ambas páginas "
                "cumplen la condición → riesgo real de doble disparo en la misma sesión.")
    ws["A7"].alignment = LEFT
    ws.row_dimensions[7].height = 45

    ws["A9"] = "3. Google Ads está optimizando sobre una conversión (Form send) que ya no es evento clave en GA4."
    ws["A9"].font = SECTION_FONT
    ws["A10"] = ("La conversión 'Form send' (importada de GA4) sigue Enabled y Primary en Ads con 163,50 "
                 "conversiones recientes, pero el evento GA4 subyacente ya no está marcado como clave. Puede "
                 "estar recibiendo datos residuales o haberse desincronizado. Verificar directamente con Google "
                 "si sigue actualizándose.")
    ws["A10"].alignment = LEFT
    ws.row_dimensions[10].height = 55

    ws["A12"] = "4. Cinco acciones de conversión 'lead' activas a la vez en la cuenta de Ads."
    ws["A12"].font = SECTION_FONT
    ws["A13"] = ("Form send, Envío de formulario (lp15/gracias), Registro (gracias/), Lead (Thank you page) y "
                 "Formulario de contacto - Enviar (Lead Ads nativo) están todas Enabled y Primary. Si más de una "
                 "está en el set de conversión usado para pujas, se está optimizando sobre una métrica inflada.")
    ws["A13"].alignment = LEFT
    ws.row_dimensions[13].height = 45

    ws["A15"] = "5. Tag 'GA PV' duplica funcionalmente el page_view automático de GA4."
    ws["A15"].font = SECTION_FONT
    ws["A16"] = "No es grave (page_view no es un evento clave), pero es ruido innecesario en el contenedor."
    ws["A16"].alignment = LEFT

    ws["A18"] = "Arquitectura objetivo propuesta"
    ws["A18"].font = SECTION_FONT
    ws["A19"] = ("Para cada conversión de negocio (ej. Lead) dejar exactamente 2 fuentes de medición redundantes, "
                 "no más: (1) un evento nativo de GA4 marcado como clave, y (2) un evento equivalente vía GTM "
                 "independiente de GA4 (o el tag de conversión de Ads directo). Así si se rompe GA4, GTM sigue "
                 "midiendo y las campañas no se quedan sin datos de optimización. Todo lo que sea una tercera, "
                 "cuarta o quinta fuente del mismo evento debe pasarse a Secundaria o eliminarse.")
    ws["A19"].alignment = LEFT
    ws.row_dimensions[19].height = 60

    ws.column_dimensions["A"].width = 120
    for row in (3, 4, 6, 7, 9, 10, 12, 13, 15, 16, 18, 19):
        ws[f"A{row}"].alignment = LEFT


def main():
    wb = openpyxl.Workbook()
    template_ws = wb.active
    template_ws.title = "Plantilla"
    build_template_sheet(template_ws)

    dcore_ws = wb.create_sheet("DCORE")
    build_dcore_sheet(dcore_ws)

    diag_ws = wb.create_sheet("Diagnóstico DCORE")
    build_diagnostico_sheet(diag_ws)

    wb.save(OUT_PATH)
    print(f"Guardado en {OUT_PATH}")


if __name__ == "__main__":
    main()
